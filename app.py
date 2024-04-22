
from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
from sqlalchemy.orm import sessionmaker

from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from conection import AlchemyConnection  # Asegúrate de que el path sea correcto

app = Flask(__name__)



@app.route('/')
def index():
    return render_template('index.html')



@app.route('/download', methods=['POST'])
def download():
    fecha_inicio = request.form['fecha_inicio']
    fecha_fin = request.form['fecha_fin']
    
    # # Establecer los parámetros de conexión
    # server = '192.168.50.58'
    # database = 'UnoEEDST'
    # username = 'sa'
    # password = 'M+xT3cn0l0g14'

    # # Crear una cadena de conexión
    # conn_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
    # Usar AlchemyConnection para establecer la conexión
    db_connection = AlchemyConnection('UNOEE')
    engine = db_connection.create_connection()
    # Session = sessionmaker(bind=engine)
    # session = Session()

    try:
        # # Conectar a la base de datos
        # conn = pyodbc.connect(conn_str)
        
        # # Crear un cursor para ejecutar consultas
        # cursor = conn.cursor()
        
        # Definir la consulta SQL dinámicamente con f-strings
        sql_query = f"""
        exec sp_pv_cons_fact_notas 4,'   ','   ','',1,0,0,0,0,'{fecha_inicio} 00:00:00','{fecha_fin} 00:00:00',7,0,0,0,0,0,0,10095,'facturas_notas_multicompañia',3224,0,1,NULL,NULL,NULL,NULL,NULL,NULL,'MAX'
        """
        # # Ejecutar la consulta SQL
        # cursor.execute(sql_query)

        # # Obtener todos los resultados
        # column_names = [column[0] for column in cursor.description]
        # rows = cursor.fetchall()
        # rows = [[str(dato) for dato in fila] for fila in rows]
        # Ejecutar la consulta y obtener los resultados directamente en un DataFrame
        df = pd.read_sql(sql_query, engine)

        # # Verificar si hay datos
        # if len(rows) < 1:  # Si hay menos de 1 filas (considerando los encabezados), no hay suficientes datos
        #     return jsonify("No hay suficientes datos para las fechas proporcionadas."), 102 
        
        # # Convertir los resultados a un DataFrame de pandas con los nombres de las columnas
        # df = pd.DataFrame(rows, columns=column_names)
        # Verificar si el DataFrame está vacío
        if df.empty:
            return jsonify("No hay suficientes datos para las fechas proporcionadas."), 102

        # Eliminar las columnas no deseadas
        columns_to_drop = ["f_rowid", "f_id_clase_docto", "f_numero", "f_cia"]
        df.drop(columns=columns_to_drop, errors='ignore', inplace=True)
        

        # Eliminar las columnas no deseadas
        columns_to_drop = ["f_rowid", "f_id_clase_docto", "f_numero", "f_cia"]
        df = df.drop(columns=columns_to_drop, errors='ignore')
        
        # Renombrar las columnas
        new_column_names = {
            "f_desc_cia": "Empresa",
            "f_tipo_docto": "Tipo de documento",
            "f_co": "Centro de operaciones",
            "f_nrodocto": "Numero de documento",
            "f_fecha": "Fecha",
            "f_usuario_creacion": "Usuario de creacion",
            "f_estado": "Estado",
            "f_usuario_aprobacion": "Usuario de aprobacion",
            "f_fecha_aprobacion": "Fecha de aprobacion",
            "f_factura_base_devol": "Factura base devolucion",
            "f_fecha_devolucion": "Fecha de devolucion",
            "f_cliente_fact_razon_soc": "Cliente razon social",
            "f_valor_subtotal": "Valor subtotal",
            "f_vendedor_razon_social": "Vendedor",
            "f_usuario_anulacion": "Usuario de anulacion",
            "f_fecha_anulacion": "Fecha de anulacion",
            "f_id_cia": "Empresa ID"
        }
        df = df.rename(columns=new_column_names)
        
        # # Guardar el DataFrame en un archivo Excel
        # df.to_excel('resultado.xlsx', index=False)
        # print("Tabla convertida exitosamente a Excel.")

        # Uso de ExcelWriter para manejar el archivo Excel
        with pd.ExcelWriter('Notas credito.xlsx', engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Resultados')
            # workbook = writer.book
            worksheet = writer.sheets['Resultados']

           # Ajustar el tamaño de las columnas automáticamente
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

            # Estilos para el encabezado
            header_font = Font(bold=True, color="ffffff")
            header_fill = PatternFill(start_color="000080", end_color="000000", fill_type="solid")
            header_alignment = Alignment(horizontal='center')
            for col_num, value in enumerate(df.columns.values):
                cell = worksheet.cell(row=1, column=col_num+1)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Crear y estilizar la tabla dentro de la hoja de cálculo
            tab = Table(displayName="Tabla1", ref=f"A1:{chr(65+len(df.columns)-1)}{len(df)+1}")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            worksheet.add_table(tab)

        return send_file('Notas credito.xlsx', as_attachment=True)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
